from sportland_scraper import SportlandScraper
from openpyxl import Workbook
from datetime import datetime

class ExcelExporter:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Products"
        
        # Set up headers
        headers = ["Nosaukums", "Zīmols", "Cena", "Produkta kods", "Pieejamie izmēri", "Saite"]
        for col, header in enumerate(headers, 1):
            self.ws.cell(row=1, column=col, value=header)
    
    def export_products(self, products, filename=None):
        if filename is None:
            # Create filename with current date
            date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sportland_products_{date_str}.xlsx"
        
        # Write product data
        for row, product in enumerate(products, 2):  # Start from row 2 (after headers)
            self.ws.cell(row=row, column=1, value=product['name'])
            self.ws.cell(row=row, column=2, value=product['brand'])
            self.ws.cell(row=row, column=3, value=product['price'])
            self.ws.cell(row=row, column=4, value=product['sku'])
            self.ws.cell(row=row, column=5, value=', '.join(product['sizes']))
            self.ws.cell(row=row, column=6, value=product['link'])
        
        # Adjust column widths
        for column in self.ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save the workbook
        self.wb.save(filename)
        return filename

def main():
    scraper = SportlandScraper()
    products = scraper.get_products("https://sportland.lv/viriesu/apavi")
    
    # Export to Excel
    exporter = ExcelExporter()
    filename = exporter.export_products(products)
    print(f"Products exported to {filename}")

if __name__ == "__main__":
    main()